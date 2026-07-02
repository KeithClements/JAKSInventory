from __future__ import annotations

import html
import io
import csv
import json
import logging
import re
from datetime import datetime
from urllib.parse import quote as url_quote

from fastapi import APIRouter, Depends, Request, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, Response
from fastapi.templating import Jinja2Templates
from sqlalchemy import func
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session, joinedload

from app.constants import (
    AddressType, CallOutcome, CallType,
    CommunicationChannel, CommunicationDirection,
    CoreStatus, PaymentTerms, Permission, PricingTier, QuoteStatus, SOStatus,
    CustomerType, CustomerStatus, CustomerFlag, CUSTOMER_TYPE_LABELS, CUSTOMER_FLAG_LABELS,
    CUSTOMER_STORED_FLAGS,
)
from app.deps import get_db, get_current_user_id
from app.models.communication import Communication
from app.models.customer import Customer, CustomerAddress, CustomerCallLog
from app.models.core import CoreCharge
from app.models.invoice import Invoice, PaymentAllocation
from app.models.quote import Quote, SalesOrder
from app.services.base import PermissionDeniedError
from app.services.crm_service import CRMService
from app.services.customer_service import (
    CustomerService, is_valid_email, normalize_phone,
)
from app.services.customer_metrics_service import CustomerMetricsService
from app.services.messaging_service import MessagingService
from app.services.quote_service import QuoteService

log = logging.getLogger(__name__)

router = APIRouter(prefix="/customers", tags=["customers"])
templates = Jinja2Templates(directory="app/templates")


def _digits(s: str | None) -> str:
    """Strip everything except digits — used to normalize phone numbers for search."""
    return "".join(ch for ch in (s or "") if ch.isdigit())


def _normalize_name(s: str | None) -> str:
    """Lowercase + strip non-alphanumerics — for fuzzy duplicate-name detection."""
    return re.sub(r"[^a-z0-9]", "", (s or "").lower())


def _form_customer_type(form) -> str:
    """Read + validate the customer_type form field, defaulting to OTHER (P2-D6).
    Keeps an unknown/blank value from ever reaching the column."""
    try:
        return CustomerType(str(form.get("customer_type", "")))
    except ValueError:
        return CustomerType.OTHER


def _form_customer_status(form) -> str:
    """Read + validate the customer_status form field, defaulting to ACTIVE."""
    try:
        return CustomerStatus(str(form.get("customer_status", "")))
    except ValueError:
        return CustomerStatus.ACTIVE


def _find_duplicate_customers(
    db: Session, company_name: str, exclude_id: int | None = None
) -> list[Customer]:
    """Return active customers whose company name looks like a duplicate of
    `company_name`. Matches on normalized exact equality, or substring either
    direction once the normalized name is long enough to be meaningful (>= 4
    chars) — this catches "Mike's Diesel" vs "Mikes Diesel Repair" without
    firing on trivially short fragments. It is a soft warning, never a block."""
    norm = _normalize_name(company_name)
    if not norm:
        return []
    matches: list[Customer] = []
    for c in db.query(Customer).filter(Customer.is_active == True).all():  # noqa: E712
        if exclude_id and c.id == exclude_id:
            continue
        cn = _normalize_name(c.company_name)
        if not cn:
            continue
        if cn == norm or (len(norm) >= 4 and len(cn) >= 4 and (cn in norm or norm in cn)):
            matches.append(c)
    return matches


def _find_customer_by_email(
    db: Session, email: str, exclude_id: int | None = None
) -> Customer | None:
    """C10 — return the existing customer that owns `email` (case-insensitive),
    if any. Unlike the company-name match this is a HARD dedup key (DB unique
    index uq_customers_email), so callers must block, not warn. Blank → no match."""
    em = (email or "").strip()
    if not em:
        return None
    q = db.query(Customer).filter(func.lower(Customer.email) == em.lower())
    if exclude_id:
        q = q.filter(Customer.id != exclude_id)
    return q.first()


def _find_customer_by_phone(
    db: Session, phone: str, exclude_id: int | None = None
) -> Customer | None:
    """QA — return an existing ACTIVE customer whose phone normalizes to the same
    digits as ``phone`` (counter staff look people up by phone, so a shared phone
    is the same fragmentation the email check prevents). Blank phone → no match
    (phoneless customers never collide). Unlike email this is NOT a unique DB key,
    so the caller WARNS with a "Create Anyway" override rather than hard-blocking.
    Delegates the normalization rule to CustomerService.find_by_phone."""
    return CustomerService(db).find_by_phone(phone, exclude_id=exclude_id)


# ── List tab definitions (JAKS_UI_Change_Plan.md §2) ─────────────────────────

# Tab slug → filter behavior.  The four operational tabs all scope to ACTIVE
# customers ("all" = every *active* customer); this is the busy counter screen
# and its default view must not surface deactivated accounts.  The "inactive"
# lifecycle tab is the only one that drops the is_active==True filter — it shows
# *only* deactivated customers so they stay reachable and can be reactivated from
# detail.  (Mirrors the Vendors fix, which defaults to `active` and hides
# deactivated records; Customers keeps `all` as its active-scoped default rather
# than a true union — see JAKS_UI_Change_Plan.md "Customers List".)
_CUST_TABS: list[tuple[str, str]] = [
    ("all",           "All"),
    ("open_invoices", "Open Invoices"),
    ("open_quotes",   "Open Quotes"),
    ("terms",         "On Terms"),
    ("inactive",      "Inactive"),   # lifecycle tab — is_active == False only
]

# Payment-terms values that fall under the "On Terms" tab.
_TERMS_VALUES = {"net_15", "net_30", "net_60"}

# Invoice statuses that count as an "open invoice" (activity chips + tab).
_OPEN_INVOICE_STATUSES = ["draft", "open", "partial"]
# Quote statuses that are terminal (NOT open) — any other status is an open quote.
_CLOSED_QUOTE_STATUSES = [QuoteStatus.CONVERTED, QuoteStatus.DECLINED, QuoteStatus.EXPIRED]


# ── List ──────────────────────────────────────────────────────────────────────

@router.get("/", response_class=HTMLResponse)
def customer_list(
    request: Request,
    q: str = "",
    tab: str = "all",
    sort: str = "company_name",
    direction: str = "asc",
    db: Session = Depends(get_db),
):
    # Normalise unknown tab slugs
    valid_tabs = {t[0] for t in _CUST_TABS}
    if tab not in valid_tabs:
        tab = "all"

    # ── Unfiltered counts ─────────────────────────────────────────────────────
    # Operational-tab counts come from the full ACTIVE dataset; the lifecycle
    # "inactive" tab is counted separately (it lives outside `all_active`).
    all_active = db.query(Customer).filter(Customer.is_active == True).all()
    all_ids = [c.id for c in all_active]
    inactive_count = (
        db.query(func.count(Customer.id)).filter(Customer.is_active == False).scalar() or 0
    )

    if all_ids:
        _inv_active = dict(
            db.query(Invoice.customer_id, func.count(Invoice.id))
            .filter(
                Invoice.customer_id.in_(all_ids),
                Invoice.status.in_(_OPEN_INVOICE_STATUSES),
            )
            .group_by(Invoice.customer_id)
            .all()
        )
        _quote_active = dict(
            db.query(Quote.customer_id, func.count(Quote.id))
            .filter(
                Quote.customer_id.in_(all_ids),
                Quote.status.notin_(_CLOSED_QUOTE_STATUSES),
            )
            .group_by(Quote.customer_id)
            .all()
        )
    else:
        _inv_active = {}
        _quote_active = {}

    counts: dict[str, int] = {
        "all":           len(all_active),
        "open_invoices": sum(1 for c in all_active if _inv_active.get(c.id, 0) > 0),
        "open_quotes":   sum(1 for c in all_active if _quote_active.get(c.id, 0) > 0),
        "terms":         sum(1 for c in all_active if (c.payment_terms or "") in _TERMS_VALUES),
        "inactive":      inactive_count,
    }

    # ── Apply tab filter ──────────────────────────────────────────────────────
    if tab == "inactive":
        # The only tab that surfaces deactivated customers — pulled with a
        # dedicated query since `all_active` (and its activity maps) exclude them.
        base_pool = db.query(Customer).filter(Customer.is_active == False).all()
    elif tab == "open_invoices":
        with_inv = {cid for cid, cnt in _inv_active.items() if cnt > 0}
        base_pool = [c for c in all_active if c.id in with_inv]
    elif tab == "open_quotes":
        with_q = {cid for cid, cnt in _quote_active.items() if cnt > 0}
        base_pool = [c for c in all_active if c.id in with_q]
    elif tab == "terms":
        base_pool = [c for c in all_active if (c.payment_terms or "") in _TERMS_VALUES]
    else:
        base_pool = all_active

    # ── Apply search ──────────────────────────────────────────────────────────
    if q:
        q_lower = q.lower()
        q_digits = _digits(q)
        customers = [
            c for c in base_pool
            if (
                (c.company_name and q_lower in c.company_name.lower())
                or (c.contact_name and q_lower in c.contact_name.lower())
                or (c.email and q_lower in c.email.lower())
                or (c.account_number and q_lower in c.account_number.lower())  # #5
                or (q_digits and c.phone and q_digits in _digits(c.phone))
            )
        ]
    else:
        customers = base_pool

    # ── Sort (#4 — whitelisted keys, asc/desc; mirrors products.py) ───────────
    _CUST_SORT_KEYS = {
        "company_name":   lambda c: (c.company_name or "").lower(),
        "account_number": lambda c: (c.account_number or "").lower(),
        "contact_name":   lambda c: (c.contact_name or "").lower(),
        "created":        lambda c: c.created_at or datetime.min,
    }
    sort = sort if sort in _CUST_SORT_KEYS else "company_name"
    direction = "desc" if str(direction).lower() == "desc" else "asc"
    customers = sorted(customers, key=_CUST_SORT_KEYS[sort], reverse=(direction == "desc"))

    # ── Per-customer activity counts (for the rows we're actually showing) ────
    customer_ids = [c.id for c in customers]
    if tab == "inactive" and customer_ids:
        # Inactive rows aren't in the active-dataset maps; look their counts up
        # directly so a deactivated account with lingering open items reads true.
        _inv_inactive = dict(
            db.query(Invoice.customer_id, func.count(Invoice.id))
            .filter(
                Invoice.customer_id.in_(customer_ids),
                Invoice.status.in_(_OPEN_INVOICE_STATUSES),
            )
            .group_by(Invoice.customer_id)
            .all()
        )
        _q_inactive = dict(
            db.query(Quote.customer_id, func.count(Quote.id))
            .filter(
                Quote.customer_id.in_(customer_ids),
                Quote.status.notin_(_CLOSED_QUOTE_STATUSES),
            )
            .group_by(Quote.customer_id)
            .all()
        )
        open_invoice_counts = {cid: _inv_inactive.get(cid, 0) for cid in customer_ids}
        open_quote_counts = {cid: _q_inactive.get(cid, 0) for cid in customer_ids}
    else:
        open_invoice_counts = {cid: _inv_active.get(cid, 0) for cid in customer_ids}
        open_quote_counts = {cid: _quote_active.get(cid, 0) for cid in customer_ids}

    # ── Last-sale dates ───────────────────────────────────────────────────────
    if customer_ids:
        _raw_dates = dict(
            db.query(Invoice.customer_id, func.max(Invoice.created_at))
            .filter(
                Invoice.customer_id.in_(customer_ids),
                Invoice.status != "void",
            )
            .group_by(Invoice.customer_id)
            .all()
        )
        last_sale_dates: dict[int, str] = {}
        for _cid, _val in _raw_dates.items():
            if _val is None:
                continue
            if hasattr(_val, "strftime"):
                last_sale_dates[_cid] = _val.strftime("%b %d")
            else:
                last_sale_dates[_cid] = str(_val)[:10]
    else:
        last_sale_dates = {}

    # ── §2B data: balance_due_map, open_so_counts, outstanding_cores_map ─────
    #
    # Scoped to `customer_ids` (the rendered subset) — never the full all_ids
    # set — so this scales with what's visible, not with the total customer count.

    if customer_ids:
        # balance_due_map — sum of Invoice.balance_due across open/partial invoices.
        # Invoice.balance_due is a Python property (total - amount_paid), so we load
        # the invoices with their lines and allocations in two eager-load queries
        # (joinedload issues one IN-query per relationship, not N+1).
        _open_invoices = (
            db.query(Invoice)
            .options(
                joinedload(Invoice.lines),
                joinedload(Invoice.allocations),
            )
            .filter(
                Invoice.customer_id.in_(customer_ids),
                Invoice.status.in_(["open", "partial"]),
            )
            .all()
        )
        balance_due_map: dict[int, float] = {}
        for _inv in _open_invoices:
            _bd = _inv.balance_due
            if _bd > 0:
                balance_due_map[_inv.customer_id] = round(
                    balance_due_map.get(_inv.customer_id, 0.0) + _bd, 2
                )

        # open_so_counts — Sales Orders in any active (not closed) status.
        # OPEN + PARTIAL + HOLD; FULFILLED/INVOICED/CANCELLED are excluded.
        _so_raw = dict(
            db.query(SalesOrder.customer_id, func.count(SalesOrder.id))
            .filter(
                SalesOrder.customer_id.in_(customer_ids),
                SalesOrder.status.in_([
                    SOStatus.OPEN,
                    SOStatus.PARTIAL,
                    SOStatus.HOLD,
                ]),
            )
            .group_by(SalesOrder.customer_id)
            .all()
        )
        open_so_counts: dict[int, int] = {
            cid: _so_raw.get(cid, 0) for cid in customer_ids
        }

        # outstanding_cores_map — CoreCharge records still open/partially returned.
        # customer_id is nullable on CoreCharge (drop-ship / no-customer cores);
        # the IS NOT NULL filter excludes those rows from the aggregate.
        _cores_raw = dict(
            db.query(CoreCharge.customer_id, func.count(CoreCharge.id))
            .filter(
                CoreCharge.customer_id.in_(customer_ids),
                CoreCharge.customer_id.isnot(None),
                CoreCharge.status.in_([CoreStatus.OPEN, CoreStatus.PARTIAL]),
            )
            .group_by(CoreCharge.customer_id)
            .all()
        )
        outstanding_cores_map: dict[int, int] = {
            cid: _cores_raw.get(cid, 0) for cid in customer_ids
        }

    else:
        balance_due_map = {}
        open_so_counts = {}
        outstanding_cores_map = {}

    # ── Phase 2 §4 — flags + relationship metrics for the visible rows ─────────
    # Scoped to `customer_ids` (the rendered subset), like the §2B maps above, so
    # it scales with what's shown. flags_for is pure-Python on already-loaded rows;
    # metrics_for_batch is the §4.4 contract (P2-Q1 single definition). The plan
    # flags cache/materialisation as the later perf path.
    if customer_ids:
        _csvc = CustomerService(db)
        customer_flags = {c.id: _csvc.flags_for(c) for c in customers}
        customer_metrics = CustomerMetricsService(db).metrics_for_batch(customer_ids)

        # last_contacted_map (#5) — most recent CustomerCallLog.logged_at per
        # visible customer.  Single grouped query, no N+1.
        _lc_raw = dict(
            db.query(
                CustomerCallLog.customer_id,
                func.max(CustomerCallLog.logged_at),
            )
            .filter(CustomerCallLog.customer_id.in_(customer_ids))
            .group_by(CustomerCallLog.customer_id)
            .all()
        )
        last_contacted_map: dict[int, datetime | None] = {
            cid: _lc_raw.get(cid) for cid in customer_ids
        }
    else:
        customer_flags = {}
        customer_metrics = {}
        last_contacted_map = {}

    return templates.TemplateResponse(
        request,
        "customers/list.html",
        {
            "customers": customers,
            "q": q,
            "tab": tab,
            "sort": sort,
            "direction": direction,
            "counts": counts,
            "tabs": _CUST_TABS,
            "open_invoice_counts": open_invoice_counts,
            "open_quote_counts": open_quote_counts,
            "last_sale_dates": last_sale_dates,
            # §2B data
            "balance_due_map": balance_due_map,
            "open_so_counts": open_so_counts,
            "outstanding_cores_map": outstanding_cores_map,
            # Phase 2 §4 contracts (UI wires columns/chips)
            "customer_flags": customer_flags,
            "customer_metrics": customer_metrics,
            "customer_flag_labels": CUSTOMER_FLAG_LABELS,
            "customer_type_labels": CUSTOMER_TYPE_LABELS,
            # #5 — last_contacted_map: {customer_id: datetime|None}
            "last_contacted_map": last_contacted_map,
        },
    )


# ── Export CSV — MUST be before /{customer_id} ───────────────────────────────

@router.get("/export.csv")
def customer_export_csv(
    tab: str = "all",
    q: str = "",
    db: Session = Depends(get_db),
):
    """
    Stream the current filtered customer list as a CSV download.
    Mirrors the list view's tab/q filters so "export what I see" matches exactly
    (same pattern as /products/export.csv).
    """
    from fastapi.responses import StreamingResponse

    valid_tabs = {t[0] for t in _CUST_TABS}
    if tab not in valid_tabs:
        tab = "all"

    if tab == "inactive":
        pool = db.query(Customer).filter(Customer.is_active == False).all()  # noqa: E712
    else:
        pool = db.query(Customer).filter(Customer.is_active == True).all()  # noqa: E712
        ids = [c.id for c in pool]
        if tab == "open_invoices" and ids:
            with_inv = {
                cid for (cid,) in db.query(Invoice.customer_id)
                .filter(
                    Invoice.customer_id.in_(ids),
                    Invoice.status.in_(_OPEN_INVOICE_STATUSES),
                )
                .distinct()
            }
            pool = [c for c in pool if c.id in with_inv]
        elif tab == "open_quotes" and ids:
            with_q = {
                cid for (cid,) in db.query(Quote.customer_id)
                .filter(
                    Quote.customer_id.in_(ids),
                    Quote.status.notin_(_CLOSED_QUOTE_STATUSES),
                )
                .distinct()
            }
            pool = [c for c in pool if c.id in with_q]
        elif tab == "terms":
            pool = [c for c in pool if (c.payment_terms or "") in _TERMS_VALUES]

    if q:
        q_lower = q.lower()
        q_digits = _digits(q)
        pool = [
            c for c in pool
            if (
                (c.company_name and q_lower in c.company_name.lower())
                or (c.contact_name and q_lower in c.contact_name.lower())
                or (c.email and q_lower in c.email.lower())
                or (c.account_number and q_lower in c.account_number.lower())
                or (q_digits and c.phone and q_digits in _digits(c.phone))
            )
        ]

    pool.sort(key=lambda c: (c.company_name or "").lower())

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([
        "company_name", "contact_name", "account_number", "phone", "email",
        "city", "state", "zip_code", "payment_terms", "pricing_tier",
        "credit_limit", "discount_pct", "is_tax_exempt", "is_active",
    ])
    for c in pool:
        writer.writerow([
            c.company_name,
            c.contact_name,
            c.account_number,
            c.phone,
            c.email,
            c.city,
            c.state,
            c.zip_code,
            c.payment_terms,
            c.pricing_tier,
            f"{c.credit_limit:.2f}" if c.credit_limit else "",
            f"{c.discount_pct:.2f}" if c.discount_pct else "",
            "yes" if c.is_tax_exempt else "no",
            "yes" if c.is_active else "no",
        ])

    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=customers.csv"},
    )


# ── Review-seeding exports ───────────────────────────────────────────────────
# Cold-start social proof: turn finalized sales history into review-request
# files. Multi-segment paths, so they never collide with /{customer_id}.
# See app/services/review_outreach_service.py for the consent + dedupe rules.

def _csv_response(filename: str, header: list[str], rows: list[list]) -> "StreamingResponse":
    from fastapi.responses import StreamingResponse
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(header)
    for r in rows:
        writer.writerow(r)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/review-requests/judgeme.csv")
def review_requests_judgeme_csv(
    paid_only: bool = True,
    db: Session = Depends(get_db),
):
    """Judge.me 'send review requests via CSV' importer file (product reviews).

    paid_only=true (default) targets only paid invoices; pass paid_only=false to
    include all finalized (open/partial/paid) sales.
    """
    from app.services.review_outreach_service import ReviewOutreachService

    rows = ReviewOutreachService(db).judgeme_rows(paid_only=paid_only)
    return _csv_response(
        "judgeme_review_requests.csv",
        ["reviewer_name", "reviewer_email", "product_id", "fulfilled_at", "quantity"],
        [[r["reviewer_name"], r["reviewer_email"], r["product_id"], r["fulfilled_at"], r["quantity"]]
         for r in rows],
    )


@router.get("/review-requests/google.csv")
def review_requests_google_csv(
    paid_only: bool = True,
    db: Session = Depends(get_db),
):
    """Personal-ask outreach list for the Google Business Profile review track.

    One row per customer (best customers first), with per-channel consent flags
    so you only email/text people who agreed to it.
    """
    from app.services.review_outreach_service import ReviewOutreachService

    rows = ReviewOutreachService(db).google_outreach_rows(paid_only=paid_only)
    return _csv_response(
        "google_review_outreach.csv",
        ["customer_name", "company_name", "email", "phone", "allow_email",
         "allow_sms", "orders", "total_spent", "last_purchase", "sample_product"],
        [[r["customer_name"], r["company_name"], r["email"], r["phone"],
          r["allow_email"], r["allow_sms"], r["orders"], r["total_spent"],
          r["last_purchase"], r["sample_product"]] for r in rows],
    )


# ── Customer preview panel (HTMX dock) ───────────────────────────────────────
# IMPORTANT: must be registered BEFORE /{customer_id} to avoid the int route
# capturing "preview" as a customer_id parameter.

@router.get("/preview/{customer_id}", response_class=HTMLResponse)
def customer_preview_panel(
    customer_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """
    Bottom-dock preview panel for the Customer List (§7 Primitive 5).
    Loaded via htmx.ajax() on row click; renders _preview_panel.html.

    Context published to UI lane:
      c                  — Customer ORM object
      open_invoice_count — int: open/draft/partial invoices for this customer
      open_quote_count   — int: active quotes
      last_sale          — str | None: formatted "Mon DD" or None
    """
    c = db.query(Customer).filter(Customer.id == customer_id).first()
    if c is None:
        return HTMLResponse(
            '<p class="px-6 py-4 text-sm text-red-500">Customer not found.</p>',
            status_code=404,
        )

    open_invoice_count = (
        db.query(func.count(Invoice.id))
        .filter(
            Invoice.customer_id == customer_id,
            Invoice.status.in_(["draft", "open", "partial"]),
        )
        .scalar()
        or 0
    )
    open_quote_count = (
        db.query(func.count(Quote.id))
        .filter(
            Quote.customer_id == customer_id,
            Quote.status.notin_([
                QuoteStatus.CONVERTED,
                QuoteStatus.DECLINED,
                QuoteStatus.EXPIRED,
            ]),
        )
        .scalar()
        or 0
    )

    _raw_last = (
        db.query(func.max(Invoice.created_at))
        .filter(Invoice.customer_id == customer_id, Invoice.status != "void")
        .scalar()
    )
    last_sale: str | None = None
    if _raw_last is not None:
        if hasattr(_raw_last, "strftime"):
            last_sale = _raw_last.strftime("%b %d")
        else:
            last_sale = str(_raw_last)[:10]

    # Balance due — load open/partial invoices with lines + allocations (3 queries, no N+1)
    _open_invs = (
        db.query(Invoice)
        .options(joinedload(Invoice.lines), joinedload(Invoice.allocations))
        .filter(
            Invoice.customer_id == customer_id,
            Invoice.status.in_(["open", "partial"]),
        )
        .all()
    )
    balance_due = round(sum(inv.balance_due for inv in _open_invs if inv.balance_due > 0), 2)

    # Open Sales Orders
    open_so_count = (
        db.query(func.count(SalesOrder.id))
        .filter(
            SalesOrder.customer_id == customer_id,
            SalesOrder.status.in_([SOStatus.OPEN, SOStatus.PARTIAL, SOStatus.HOLD]),
        )
        .scalar()
        or 0
    )

    # Outstanding core charges
    outstanding_core_count = (
        db.query(func.count(CoreCharge.id))
        .filter(
            CoreCharge.customer_id == customer_id,
            CoreCharge.status.in_([CoreStatus.OPEN, CoreStatus.PARTIAL]),
        )
        .scalar()
        or 0
    )

    # last_contact — most recent CustomerCallLog.logged_at.
    # Delegated to CustomerService.last_contacted so the query lives in exactly
    # one place (single-customer path; list route uses a grouped query instead).
    _csvc = CustomerService(db)
    last_contact = _csvc.last_contacted(customer_id)  # datetime | None
    return templates.TemplateResponse(
        request,
        "customers/_preview_panel.html",
        {
            "c": c,
            "open_invoice_count": open_invoice_count,
            "open_quote_count": open_quote_count,
            "last_sale": last_sale,
            "last_contact": last_contact,  # Seam 2 — for the dynamic preview dock
            "balance_due": balance_due,
            "open_so_count": open_so_count,
            "outstanding_core_count": outstanding_core_count,
            # Phase 2 §4 contracts (condensed on preview)
            "flags": _csvc.flags_for(c),
            "metrics": CustomerMetricsService(db).metrics_for(c),
            "credit_status": _csvc.credit_status(c),
            "customer_flag_labels": CUSTOMER_FLAG_LABELS,
            "customer_type_labels": CUSTOMER_TYPE_LABELS,
        },
    )


# ── New ───────────────────────────────────────────────────────────────────────

def _new_customer_ctx(db: Session) -> dict:
    """Shared context for the new-customer form (GET and the dup-warning
    re-render both need it, or the Customer Type select + type-defaults pre-fill
    JSON go missing on the dup path)."""
    return {
        "payment_terms": list(PaymentTerms),
        "pricing_tiers": list(PricingTier),
        # P2-D1/D6 — Customer Type + type-driven default profiles. The UI picks
        # a type, then pre-fills the form from `type_defaults[<type>]` client-side.
        "customer_types": list(CustomerType),
        "customer_type_labels": CUSTOMER_TYPE_LABELS,
        "customer_flag_labels": CUSTOMER_FLAG_LABELS,
        "type_defaults": CustomerService(db).all_type_defaults(),
    }


@router.get("/new", response_class=HTMLResponse)
def customer_new(request: Request, db: Session = Depends(get_db)):
    return templates.TemplateResponse(request, "customers/new.html", _new_customer_ctx(db))


@router.get("/type-defaults/{customer_type}")
def customer_type_defaults(customer_type: str, db: Session = Depends(get_db)):
    """JSON contract — resolved default profile for one Customer Type. The
    new-customer form can fetch this on type-change as an alternative to the
    embedded `type_defaults` map. Unknown types resolve to the OTHER profile."""
    return CustomerService(db).resolve_defaults(customer_type)


@router.post("/new", response_class=RedirectResponse)
async def customer_create(
    request: Request,
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user_id),
):
    form = await request.form()
    company_name = str(form.get("company_name", "")).strip()
    email = str(form.get("email", "")).strip()
    phone = str(form.get("phone", "")).strip()

    # QA — server-side email-format validation. The browser only enforces
    # type=email; a scripted/old-browser POST of "notanemail" otherwise saves.
    # Blank stays allowed (matches the column default ''); only validate a value.
    if email and not is_valid_email(email):
        return templates.TemplateResponse(
            request,
            "customers/new.html",
            {
                **_new_customer_ctx(db),
                "email_error": "Please enter a valid email address (e.g. name@example.com).",
                "prefill": {k: str(v) for k, v in form.items()},
            },
            status_code=422,
        )

    # A standing discount outside 0–100 poisons every downstream document header:
    # quote/invoice services now REJECT bad values, so a bad stored customer
    # default would brick "+ Quote" / "New Invoice" for that customer entirely.
    try:
        _discount = float(form.get("discount_pct") or 0)
    except (TypeError, ValueError):
        _discount = -1.0
    if not (0.0 <= _discount <= 100.0):
        return templates.TemplateResponse(
            request,
            "customers/new.html",
            {
                **_new_customer_ctx(db),
                "email_error": "Standing discount must be between 0 and 100.",
                "prefill": {k: str(v) for k, v in form.items()},
            },
            status_code=422,
        )

    # C10 — HARD email dedup. A duplicate email means split AR / inconsistent
    # credit hold / duplicate QBO push, so this is a block (not a "create
    # anyway" warning like the company-name case). Pre-check for a friendly
    # message; the DB unique index is the backstop below if a race slips past.
    email_conflict = _find_customer_by_email(db, email)
    if email_conflict is not None:
        return templates.TemplateResponse(
            request,
            "customers/new.html",
            {
                **_new_customer_ctx(db),
                "email_conflict": email_conflict,
                "prefill": {k: str(v) for k, v in form.items()},
            },
        )

    # Duplicate protection — warn instead of silently creating a near-duplicate,
    # unless the user explicitly chose "Create Anyway" (confirm_duplicate=1).
    # Covers BOTH a similar company name AND a shared phone (counter staff look
    # customers up by phone, so a duplicate phone fragments AR exactly like a
    # duplicate name). Phone is not a unique DB key, so it warns (override-able)
    # rather than hard-blocking like email.
    if str(form.get("confirm_duplicate", "")) != "1":
        dup_matches = _find_duplicate_customers(db, company_name) if company_name else []
        phone_conflict = _find_customer_by_phone(db, phone)
        if dup_matches or phone_conflict is not None:
            return templates.TemplateResponse(
                request,
                "customers/new.html",
                {
                    **_new_customer_ctx(db),
                    "dup_matches": dup_matches,
                    "phone_conflict": phone_conflict,
                    "prefill": {k: str(v) for k, v in form.items()},
                },
            )

    # blank → NULL (use system default); explicit value (incl. 0) overrides — mirrors update
    _cs = str(form.get("card_surcharge_pct", "")).strip()
    c = Customer(
        company_name=str(form.get("company_name", "")).strip(),
        contact_name=str(form.get("contact_name", "")).strip(),
        customer_type=_form_customer_type(form),
        customer_status=_form_customer_status(form),
        account_number=str(form.get("account_number", "")).strip(),
        card_surcharge_pct=float(_cs) if _cs else None,
        phone=str(form.get("phone", "")).strip(),
        email=str(form.get("email", "")).strip(),
        address_line1=str(form.get("address_line1", "")).strip(),
        address_line2=str(form.get("address_line2", "")).strip(),
        city=str(form.get("city", "")).strip(),
        state=str(form.get("state", "")).strip(),
        zip_code=str(form.get("zip_code", "")).strip(),
        payment_terms=str(form.get("payment_terms", PaymentTerms.COD)),
        pricing_tier=str(form.get("pricing_tier", "standard")),
        credit_limit=float(form.get("credit_limit") or 0),
        discount_pct=float(form.get("discount_pct") or 0),
        interest_rate=float(form.get("interest_rate") or 0),
        is_tax_exempt=bool(form.get("is_tax_exempt")),
        tax_exempt_cert_number=str(form.get("tax_exempt_cert_number", "")).strip() or None,
        notes=str(form.get("notes", "")).strip(),
        internal_notes=str(form.get("internal_notes", "")).strip(),
    )
    db.add(c)
    # P2-D2 — flags chip editor (Requires-PO / Credit-Hold / Call-first /
    # Warranty-escalation). No-op for a new customer when the form omits them.
    CustomerService(db).set_stored_flags(c, form.getlist("flags"))
    try:
        db.commit()
    except IntegrityError:
        # C10 backstop — a concurrent insert won the email race (or another
        # unique key collided). Re-render with the conflict instead of a 500.
        db.rollback()
        return templates.TemplateResponse(
            request,
            "customers/new.html",
            {
                **_new_customer_ctx(db),
                "email_conflict": _find_customer_by_email(db, email),
                "prefill": {k: str(v) for k, v in form.items()},
            },
        )
    return RedirectResponse(f"/customers/{c.id}", status_code=303)


# ── Quick Create (slide-over — called from quote/invoice customer field [+]) ──

@router.get("/quick-create-form", response_class=HTMLResponse)
def customer_quick_create_form(request: Request):
    return templates.TemplateResponse(request, "customers/_quick_create.html")


@router.post("/quick-create", response_class=HTMLResponse)
async def customer_quick_create(request: Request, db: Session = Depends(get_db), user_id: int = Depends(get_current_user_id)):
    form = await request.form()
    company_name = str(form.get("company_name", "")).strip()
    if not company_name:
        return HTMLResponse(
            '<p class="text-sm text-red-600 font-medium px-5 py-3">Company name is required.</p>',
            status_code=422,
        )

    _action = str(form.get("_action", "save"))

    # QA — server-side email-format validation (parity with the full form). The
    # quick-create email field is type=email only client-side; blank stays OK.
    _qc_email = str(form.get("email", "")).strip()
    if _qc_email and not is_valid_email(_qc_email):
        return HTMLResponse(
            '<p class="text-sm text-red-600 font-medium px-5 py-3">'
            'Please enter a valid email address (e.g. name@example.com).</p>',
            status_code=422,
        )

    # ── Duplicate protection ──────────────────────────────────────────────────
    # Unless the user explicitly chose "Create Anyway" (confirm_duplicate=1), warn
    # about existing/similar company names OR a shared phone instead of silently
    # creating a dupe. The phone match is folded into the same dup_matches list
    # the warning banner + "Create Anyway" button already render (a counter person
    # looks customers up by phone, so a shared phone fragments AR like a dup name).
    confirm_duplicate = str(form.get("confirm_duplicate", "")) == "1"
    if not confirm_duplicate:
        dup_matches = _find_duplicate_customers(db, company_name)
        phone_conflict = _find_customer_by_phone(db, str(form.get("phone", "")).strip())
        if phone_conflict is not None and not any(m.id == phone_conflict.id for m in dup_matches):
            dup_matches = [*dup_matches, phone_conflict]
        if dup_matches:
            return templates.TemplateResponse(
                request,
                "customers/_quick_create.html",
                {
                    "dup_matches": dup_matches,
                    "prefill": {k: str(v) for k, v in form.items()},
                },
            )

    # Checkboxes: absent = unchecked, "on" = checked
    is_tax_exempt = form.get("is_tax_exempt") is not None
    ship_same     = form.get("ship_same") is not None

    c = Customer(
        company_name=company_name,
        contact_name=str(form.get("contact_name", "")).strip(),
        phone=str(form.get("phone", "")).strip(),
        email=str(form.get("email", "")).strip(),
        payment_terms=str(form.get("payment_terms", PaymentTerms.NET_30)),
        pricing_tier=str(form.get("pricing_tier", "standard")),
        credit_limit=float(form.get("credit_limit") or 0),
        discount_pct=float(form.get("discount_pct") or 0),
        is_tax_exempt=is_tax_exempt,
        tax_exempt_cert_number=str(form.get("tax_exempt_cert_number", "")).strip() or None,
        address_line1=str(form.get("address_line1", "")).strip(),
        address_line2=str(form.get("address_line2", "")).strip(),
        city=str(form.get("city", "")).strip(),
        state=str(form.get("state", "")).strip(),
        zip_code=str(form.get("zip_code", "")).strip(),
        notes=str(form.get("notes", "")).strip(),
    )
    db.add(c)
    db.flush()  # populate c.id before creating child records

    # Create a separate shipping address when the user unchecked "Same as billing"
    if not ship_same:
        ship_street = str(form.get("ship_address_line1", "")).strip()
        if ship_street:  # only if user actually filled something in
            ship_addr = CustomerAddress(
                customer_id=c.id,
                address_type=AddressType.SHIPPING,
                is_default_shipping=True,
                street=ship_street,
                street_line2=str(form.get("ship_address_line2", "")).strip(),
                city=str(form.get("ship_city", "")).strip(),
                state=str(form.get("ship_state", "")).strip(),
                zip_code=str(form.get("ship_zip_code", "")).strip(),
            )
            db.add(ship_addr)

    db.commit()

    # ── Save & New: return fresh form with in-panel success flash ──────────
    if _action == "save_new":
        return templates.TemplateResponse(
            request,
            "customers/_quick_create.html",
            {"success_flash": f"✓ {company_name} saved."},
        )

    # ── Save & Quote: create a draft quote and drop into the quote workspace ──
    # Mirrors POST /quotes/new (row "+ Quote" + preview-panel Quote both do this),
    # so the "Save & Quote →" label actually delivers the quote screen.
    if _action == "save_quote":
        quote = QuoteService(db, user_id).create_quote(
            customer_id=c.id,
            data={
                "discount_pct": c.discount_pct,
                "validity_days": 30,
                "notes": "",
            },
        )
        db.commit()
        return HTMLResponse(
            "<span></span>",
            headers={"HX-Redirect": f"/quotes/{quote.id}"},
        )

    # ── Default (Save Customer): fire record-created + show toast ──────────
    _detail = html.escape(json.dumps({"type": "customer", "id": c.id, "label": c.company_name}))
    _name   = html.escape(c.company_name)
    return HTMLResponse(
        f"""<span></span>
<div id="toast-container" hx-swap-oob="beforeend">
  <div x-data x-init="
      setTimeout(() => $el.remove(), 4000);
      window.dispatchEvent(new CustomEvent('record-created', {{ detail: {_detail} }}));
    "
    class="toast toast-success">
    <svg class="w-4 h-4 shrink-0" fill="currentColor" viewBox="0 0 20 20">
      <path fill-rule="evenodd"
            d="M10 18a8 8 0 100-16 8 8 0 000 16zm3.707-9.293a1 1 0 00-1.414-1.414L9 10.586 7.707 9.293a1 1 0 00-1.414 1.414l2 2a1 1 0 001.414 0l4-4z"
            clip-rule="evenodd"/>
    </svg>
    Customer created: {_name}
  </div>
</div>"""
    )


# ── Global typeahead search (used by Log Call slide-over in base.html) ────────

@router.get("/search-json", response_class=HTMLResponse)
def customer_search_partial(request: Request, q: str = "", db: Session = Depends(get_db)):
    """Returns an HTML partial — list of customers matching q for typeahead dropdowns.

    Searches: company_name, contact_name, phone, email (all case-insensitive ilike).
    Requires q >= 2 chars. Returns up to 8 active customers ordered by company name.
    Bug fix 2026-05-29: added contact_name and email to the OR filter (previously
    only company_name | phone, causing contact-name-only searches to return nothing).
    """
    customers: list[Customer] = []
    if q and len(q) >= 2:
        like = f"%{q}%"
        customers = (
            db.query(Customer)
            .filter(
                Customer.is_active == True,
                (
                    Customer.company_name.ilike(like)
                    | Customer.contact_name.ilike(like)
                    | Customer.phone.ilike(like)
                    | Customer.email.ilike(like)
                ),
            )
            .order_by(Customer.company_name)
            .limit(8)
            .all()
        )
    return templates.TemplateResponse(
        request,
        "customers/_search_results.html",
        {"customers": customers},
    )


# ── Global Log Call (from header button — customer selected in form body) ─────

@router.post("/log-call", response_class=HTMLResponse)
async def log_call_global(request: Request, db: Session = Depends(get_db), user_id: int = Depends(get_current_user_id)):
    """
    Handles the Quick Log Call slide-over in base.html.
    customer_id comes from the form body (chosen via typeahead, not URL).
    Returns OOB toast on success; error HTML on validation failure.
    """
    form = await request.form()
    customer_id_raw = form.get("customer_id", "")
    if not customer_id_raw:
        return HTMLResponse(
            '<p class="text-sm text-red-600 font-medium">Please select a customer before saving.</p>',
            status_code=422,
        )

    try:
        customer_id_int = int(customer_id_raw)
    except ValueError:
        return HTMLResponse(
            '<p class="text-sm text-red-600 font-medium">Invalid customer selection.</p>',
            status_code=422,
        )

    customer = db.query(Customer).filter(Customer.id == customer_id_int).first()
    if not customer:
        return HTMLResponse(
            '<p class="text-sm text-red-600 font-medium">Customer not found.</p>',
            status_code=422,
        )

    call_type = str(form.get("call_type", CallType.INBOUND))
    outcome   = str(form.get("outcome", CallOutcome.OTHER))
    notes     = str(form.get("notes", "")).strip()

    crm = CRMService(db, current_user_id=user_id)
    crm.log_call(
        customer_id=customer.id,
        call_type=call_type,
        outcome=outcome,
        notes=notes,
    )

    # Primary swap target is #log-call-result (empty on success).
    # OOB swap appends a toast to #toast-container.
    return HTMLResponse(
        f"""<span></span>
<div id="toast-container" hx-swap-oob="beforeend">
  <div x-data x-init="setTimeout(() => $el.remove(), 4000)"
       class="toast toast-success">
    <svg class="w-4 h-4 shrink-0" fill="currentColor" viewBox="0 0 20 20">
      <path fill-rule="evenodd"
            d="M10 18a8 8 0 100-16 8 8 0 000 16zm3.707-9.293a1 1 0 00-1.414-1.414L9 10.586 7.707 9.293a1 1 0 00-1.414 1.414l2 2a1 1 0 001.414 0l4-4z"
            clip-rule="evenodd"/>
    </svg>
    Call logged for {html.escape(customer.company_name)}.
  </div>
</div>"""
    )


# ── Excel / CSV Import ────────────────────────────────────────────────────────
# IMPORTANT: These routes MUST remain before /{customer_id} (the wildcard).
# FastAPI matches routes in declaration order; if /{customer_id} came first,
# "import" would be captured as a customer_id, fail int conversion, and 422.

# Maps common column name variants → canonical field name
_IMPORT_ALIASES: dict[str, str] = {
    # ── Company name (§1.2h fix: add CRM export variants) ─────────────────────
    "company": "company_name",
    "company name": "company_name",
    "company_name": "company_name",
    "business": "company_name",
    "business name": "company_name",
    # QuickBooks exports "Customer Name"; Salesforce/HubSpot use "Account Name";
    # Sage/Xero use "Client" or "Client Name"; generic exports use "Name" too.
    "customer name": "company_name",
    "customer": "company_name",
    "account name": "company_name",
    "account": "company_name",
    "client name": "company_name",
    "client": "company_name",
    "vendor name": "company_name",   # vendors imported as customers occasionally
    "organization": "company_name",
    "organisation": "company_name",
    "org": "company_name",
    # ── Contact name ──────────────────────────────────────────────────────────
    "contact": "contact_name",
    "contact name": "contact_name",
    "name": "contact_name",
    # ── Phone (§1.2h fix: mobile number + office variants) ───────────────────
    "phone": "phone",
    "phone number": "phone",
    "tel": "phone",
    "telephone": "phone",
    "mobile": "phone",
    "mobile phone": "phone",
    "mobile number": "phone",       # missing — common CRM export
    "cell": "phone",
    "cell phone": "phone",
    "work phone": "phone",
    "office phone": "phone",        # missing — common CRM export
    "direct": "phone",              # missing — common CRM export
    "main phone": "phone",
    "primary phone": "phone",
    "phone 1": "phone",
    "phone1": "phone",
    "ph": "phone",
    "phone #": "phone",
    # ── Email ─────────────────────────────────────────────────────────────────
    "email": "email",
    "email address": "email",
    "e-mail": "email",
    "e-mail address": "email",
    "email_address": "email",
    "work email": "email",
    "primary email": "email",
    "business email": "email",
    "email1": "email",
    "email 1": "email",
    "e mail": "email",
    "address": "address_line1",
    "address line 1": "address_line1",
    "address_line1": "address_line1",
    "address2": "address_line2",
    "address line 2": "address_line2",
    "address_line2": "address_line2",
    "city": "city",
    "state": "state",
    "st": "state",
    "zip": "zip_code",
    "zip code": "zip_code",
    "postal": "zip_code",
    "postal code": "zip_code",
    "zip_code": "zip_code",
    "terms": "payment_terms",
    "payment terms": "payment_terms",
    "payment_terms": "payment_terms",
    "discount": "discount_pct",
    "discount %": "discount_pct",
    "discount_pct": "discount_pct",
    "interest": "interest_rate",
    "interest rate": "interest_rate",
    "interest rate %": "interest_rate",
    "interest %": "interest_rate",
    "interest_rate": "interest_rate",
    "notes": "notes",
    "note": "notes",
    "internal notes": "internal_notes",
    "internal_notes": "internal_notes",
    "is_taxable": "is_taxable",
    "taxable": "is_taxable",
    "tax exempt": "is_tax_exempt",
    "is_tax_exempt": "is_tax_exempt",
}

_VALID_TERMS = {t.value for t in PaymentTerms}


def _normalise_terms(raw: str) -> str:
    """Map free-text payment terms to a PaymentTerms value, default COD."""
    raw = raw.strip().lower().replace(" ", "_").replace("-", "_")
    if raw in _VALID_TERMS:
        return raw
    if "30" in raw:
        return PaymentTerms.NET_30
    if "60" in raw:
        return PaymentTerms.NET_60
    return PaymentTerms.COD


def _safe_float(val: object, default: float = 0.0) -> float:
    """Parse a float from an import cell; tolerates '5%', '$10', '1,200' etc."""
    try:
        return float(str(val).strip().replace("%", "").replace("$", "").replace(",", "") or default)
    except (ValueError, TypeError):
        return default


def _parse_rows(raw_rows: list[dict[str, str]]) -> tuple[list[dict], list[dict]]:
    """
    Convert raw header→value dicts into canonical Customer field dicts.
    Returns (valid_rows, skipped_rows).
    Each skipped row has an extra 'error' key.
    """
    valid: list[dict] = []
    skipped: list[dict] = []

    for i, row in enumerate(raw_rows, start=2):  # row 2 = first data row after header
        # Build normalised field dict
        canonical: dict[str, str] = {}
        for col_raw, val in row.items():
            key = _IMPORT_ALIASES.get(col_raw.strip().lower())
            if key:
                canonical[key] = val.strip()

        company = canonical.get("company_name", "").strip()
        if not company:
            skipped.append({**row, "error": f"Row {i}: missing company name"})
            continue

        # Resolve is_tax_exempt — accept both is_tax_exempt and is_taxable columns.
        # is_taxable is the inverse of is_tax_exempt.
        is_tax_exempt_raw = canonical.get("is_tax_exempt", "")
        is_taxable_raw = canonical.get("is_taxable", "")
        if is_tax_exempt_raw:
            is_tax_exempt = is_tax_exempt_raw.lower() in ("true", "yes", "1", "y")
        elif is_taxable_raw:
            is_tax_exempt = is_taxable_raw.lower() not in ("true", "yes", "1", "y")
        else:
            is_tax_exempt = False

        valid.append({
            "company_name":   company,
            "contact_name":   canonical.get("contact_name", ""),
            "phone":          canonical.get("phone", ""),
            "email":          canonical.get("email", ""),
            "address_line1":  canonical.get("address_line1", ""),
            "address_line2":  canonical.get("address_line2", ""),
            "city":           canonical.get("city", ""),
            "state":          canonical.get("state", ""),
            "zip_code":       canonical.get("zip_code", ""),
            "payment_terms":  _normalise_terms(canonical.get("payment_terms", "")),
            "discount_pct":   _safe_float(canonical.get("discount_pct", "0")),
            "interest_rate":  _safe_float(canonical.get("interest_rate", "0")),
            "is_tax_exempt":  is_tax_exempt,
            "notes":          canonical.get("notes", ""),
            "internal_notes": canonical.get("internal_notes", ""),
        })

    return valid, skipped


def _read_xlsx(content: bytes) -> list[dict[str, str]]:
    from openpyxl import load_workbook
    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h or "").strip() for h in rows[0]]
    result = []
    for row in rows[1:]:
        if all((v is None or str(v).strip() == "") for v in row):
            continue  # skip blank rows
        result.append({headers[i]: str(v or "").strip() for i, v in enumerate(row) if i < len(headers)})
    wb.close()
    return result


def _read_csv(content: bytes) -> list[dict[str, str]]:
    try:
        text = content.decode("utf-8-sig")  # handles BOM from Excel CSV exports
    except UnicodeDecodeError:
        text = content.decode("cp1252", errors="replace")  # Excel "CSV (Comma delimited)" saves ANSI
    # newline="" lets the csv module handle \r\n / bare-\r endings and embedded
    # newlines inside quoted fields (QBO multi-line billing addresses) itself.
    reader = csv.DictReader(io.StringIO(text, newline=""))
    rows: list[dict[str, str]] = []
    for row in reader:
        cleaned = {
            (k or "").strip(): " ".join(str(v or "").split())
            for k, v in row.items()
            if k is not None and not isinstance(v, list)
        }
        if any(cleaned.values()):
            rows.append(cleaned)
    return rows


@router.get("/import", response_class=HTMLResponse)
def customer_import_form(request: Request):
    return templates.TemplateResponse(
        request,
        "customers/import.html",
        {
            "preview_rows": None,
            "import_json": None,
            "skipped": [],
            "total_valid": 0,
            "total_skipped": 0,
        },
    )


@router.post("/import", response_class=HTMLResponse)
async def customer_import_preview(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """Parse the uploaded file and return a preview — no DB writes yet."""
    content = await file.read()
    filename = (file.filename or "").lower()

    try:
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            raw_rows = _read_xlsx(content)
        elif filename.endswith(".csv"):
            raw_rows = _read_csv(content)
        else:
            return templates.TemplateResponse(
                request,
                "customers/import.html",
                {
                    "preview_rows": None,
                    "import_json": None,
                    "skipped": [],
                    "total_valid": 0,
                    "total_skipped": 0,
                    "error": "Unsupported file type. Please upload a .xls, .xlsx, or .csv file.",
                },
            )
    except Exception as exc:
        log.exception("Failed to parse import file %s", filename)
        return templates.TemplateResponse(
            request,
            "customers/import.html",
            {
                "preview_rows": None,
                "import_json": None,
                "skipped": [],
                "total_valid": 0,
                "total_skipped": 0,
                "error": f"Could not read file: {exc}",
            },
        )

    valid_rows, skipped_rows = _parse_rows(raw_rows)

    # Flag existing duplicates (by company name, case-insensitive, active customers only).
    existing_names = {
        r[0].lower()
        for r in db.query(Customer.company_name).filter(Customer.is_active == True).all()  # noqa: E712
    }
    for row in valid_rows:
        row["_duplicate"] = row["company_name"].lower() in existing_names

    return templates.TemplateResponse(
        request,
        "customers/import.html",
        {
            "preview_rows": valid_rows[:200],
            "import_json": json.dumps(valid_rows),
            "skipped": skipped_rows,
            "total_valid": len(valid_rows),
            "total_skipped": len(skipped_rows),
            "error": None,
        },
    )


@router.post("/import/confirm", response_class=RedirectResponse)
async def customer_import_confirm(
    request: Request,
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user_id),
):
    """Write validated rows to DB.

    Gated by IMPORT_CUSTOMERS — a bulk write of the customer master must not be
    reachable by a counter clerk when the single-create path is role-gated.

    Per-row guard rails mirror the single-create handler (POST /customers/new):
      • email format (is_valid_email) — a malformed email is skipped, never saved.
      • HARD email dedup (uq_customers_email) — applies regardless of
        skip_duplicates, exactly like the single-create block (no override).
      • name dedup honours the skip_duplicates checkbox: lowercase-exact against
        ALL customers (including inactive, so soft-deleted records are not
        silently re-created) PLUS the same fuzzy match the single-create warn
        uses (_find_duplicate_customers rule, active customers only). Unchecking
        skip_duplicates is the import's "Create Anyway".

    Rows are committed INDIVIDUALLY so one poisoned row (e.g. a constraint hit
    the pre-checks could not predict) skips that row only — it can no longer
    roll back the whole batch. Every skip carries a reason surfaced in the
    redirect summary.
    """
    try:
        CustomerService(db, user_id).assert_can(Permission.IMPORT_CUSTOMERS)
    except PermissionDeniedError:
        return RedirectResponse(
            "/customers/import?error=You+do+not+have+permission+to+import+customers.",
            status_code=303,
        )

    form = await request.form()
    import_json = str(form.get("import_json", "")).strip()
    skip_dupes = str(form.get("skip_duplicates", "1")) == "1"

    try:
        rows = json.loads(import_json)
        if not isinstance(rows, list):
            raise ValueError("payload must be a list")
    except (json.JSONDecodeError, ValueError):
        return RedirectResponse("/customers/import?error=Invalid+import+data", status_code=303)

    # Dedup keys are prefetched once and extended as rows commit, so a duplicate
    # WITHIN the batch is caught the same way as one already in the DB.
    existing_names: set[str] = set()
    existing_norm_names: set[str] = set()
    if skip_dupes:
        existing_names = {
            r[0].lower()
            for r in db.query(Customer.company_name).all()
        }
        existing_norm_names = {
            n for n in (
                _normalize_name(r[0])
                for r in db.query(Customer.company_name)
                .filter(Customer.is_active == True).all()  # noqa: E712
            ) if n
        }
    # Email is checked against ALL customers (the unique index has no is_active
    # carve-out) and independently of skip_duplicates — it is a hard key.
    existing_emails = {
        r[0].lower()
        for r in db.query(Customer.email).filter(Customer.email != "").all()
    }

    # Import dedup is EXACT-normalized-name only. The interactive single-create
    # flow also warns on substring matches, but there the operator can override;
    # an import has no per-row override, and the substring rule silently drops
    # distinct companies ('Cummins Northwest' skipped because 'Cummins' exists).
    def _is_name_dup(norm: str) -> bool:
        return bool(norm) and norm in existing_norm_names

    created = 0
    skip_reasons: dict[str, int] = {}

    def _skip(reason: str) -> None:
        skip_reasons[reason] = skip_reasons.get(reason, 0) + 1

    for row in rows:
        if not isinstance(row, dict):
            continue
        company = str(row.get("company_name", "") or "").strip()
        if not company:
            continue
        email = str(row.get("email", "") or "").strip()

        if email and not is_valid_email(email):
            _skip("invalid email")
            continue
        if email and email.lower() in existing_emails:
            _skip("duplicate email")
            continue
        if skip_dupes and (
            company.lower() in existing_names
            or _is_name_dup(_normalize_name(company))
        ):
            _skip("duplicate")
            continue
        # Out-of-range standing discount would brick quote/invoice creation for
        # this customer (document services reject it) — skip with a named reason.
        _discount = _safe_float(row.get("discount_pct", 0))
        if not (0.0 <= _discount <= 100.0):
            _skip("invalid discount")
            continue

        db.add(Customer(
            company_name=company,
            contact_name=row.get("contact_name", ""),
            phone=row.get("phone", ""),
            email=email,
            address_line1=row.get("address_line1", ""),
            address_line2=row.get("address_line2", ""),
            city=row.get("city", ""),
            state=row.get("state", ""),
            zip_code=row.get("zip_code", ""),
            payment_terms=row.get("payment_terms", PaymentTerms.COD),
            discount_pct=_discount,
            interest_rate=_safe_float(row.get("interest_rate", 0)),
            is_tax_exempt=bool(row.get("is_tax_exempt", False)),
            notes=row.get("notes", ""),
            internal_notes=row.get("internal_notes", ""),
        ))
        try:
            db.commit()
        except Exception:
            db.rollback()
            log.exception("Import failed writing customer %r", company)
            _skip("database error")
            continue

        created += 1
        existing_names.add(company.lower())
        norm = _normalize_name(company)
        if norm:
            existing_norm_names.add(norm)
        if email:
            existing_emails.add(email.lower())

    skipped = sum(skip_reasons.values())
    msg = f"Imported {created} customer{'s' if created != 1 else ''}"
    if skipped:
        detail = ", ".join(f"{n} {reason}" for reason, n in skip_reasons.items())
        msg += f" (skipped {skipped}: {detail})"
    return RedirectResponse(f"/customers/?ok={url_quote(msg)}", status_code=303)


# ── Balance mini-panel (HTMX partial for Quote / Invoice workspace headers) ───

@router.get("/{customer_id}/balance-mini", response_class=HTMLResponse)
def customer_balance_mini(
    customer_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """
    Lightweight partial: balance chips used in Quote and Invoice workspace headers.
    Called via hx-get with hx-trigger="load" so it loads after the page renders.
    """
    from app.services.statement_service import StatementService
    summary = StatementService(db).get_customer_balance_summary(customer_id)
    return templates.TemplateResponse(
        request,
        "customers/_balance_mini.html",
        {**summary},
    )


# ── Detail ────────────────────────────────────────────────────────────────────

@router.get("/{customer_id}", response_class=HTMLResponse)
def customer_detail(
    customer_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    c = db.query(Customer).filter(Customer.id == customer_id).first()
    if not c:
        return RedirectResponse("/customers/", status_code=303)

    recent_invoices = (
        db.query(Invoice)
        .filter(
            Invoice.customer_id == customer_id,
            Invoice.status != "void",
        )
        .order_by(Invoice.created_at.desc())
        .limit(10)
        .all()
    )

    open_quotes = (
        db.query(Quote)
        .filter(
            Quote.customer_id == customer_id,
            # Phase-1.1 — use the canonical closed-status list (includes EXPIRED)
            # so expired quotes don't inflate the badge or the Open-quotes panel,
            # matching the list/preview paths.
            Quote.status.notin_(_CLOSED_QUOTE_STATUSES),
        )
        .order_by(Quote.created_at.desc())
        .all()
    )

    svc = CustomerService(db)

    # ── Customer-Specific Product Pricing — active deals for the §"Pricing &
    #    Deals" panel (pricing_deals_panel) + the Quick Deal modal options. ─────
    # Shared with the HTMX panel-refresh routes via build_customer_rules_context
    # (single source of truth for the per-rule dicts + scope-picker options).
    from app.routers.pricing_rules import build_customer_rules_context

    _deals_ctx = build_customer_rules_context(db, c)
    pricing_rules = _deals_ctx["rules"]
    deal_categories = _deals_ctx["deal_categories"]
    deal_brands = _deals_ctx["deal_brands"]

    return templates.TemplateResponse(
        request,
        "customers/detail.html",
        {
            "customer": c,
            "recent_invoices": recent_invoices,
            "open_quotes": open_quotes,
            "payment_terms": list(PaymentTerms),
            "pricing_tiers": list(PricingTier),
            "call_types": list(CallType),
            "call_outcomes": list(CallOutcome),
            # ── Customer-Specific Product Pricing (pricing_deals_panel + modal) ──
            "rules": pricing_rules,
            "customer_id": c.id,
            "deal_categories": deal_categories,
            "deal_brands": deal_brands,
            # ── Phase 2 §4 contracts (UI wires the chips/panel) ───────────────
            "customer_types": list(CustomerType),
            "customer_type_labels": CUSTOMER_TYPE_LABELS,
            "customer_flag_labels": CUSTOMER_FLAG_LABELS,
            "flags": svc.flags_for(c),                       # §4.3 chips (merged view)
            # Edit-form flag chip editor (operator-settable flags only) + the
            # customer's currently-stored set so checkboxes reflect state.
            "stored_flag_options": [
                (f, CUSTOMER_FLAG_LABELS.get(f, f)) for f in (
                    CustomerFlag.REQUIRES_PO, CustomerFlag.CALL_FIRST,
                    CustomerFlag.WARRANTY_ESCALATION, CustomerFlag.CREDIT_HOLD,
                ) if f in CUSTOMER_STORED_FLAGS
            ],
            "customer_stored_flags": set(CustomerService._parse_stored(c.flags)),
            "metrics": CustomerMetricsService(db).metrics_for(c),  # §4.4 live panel
            "credit_status": svc.credit_status(c),           # §4.5 warn-only
            # §4.6 / #8 unified timeline — unblocks _timeline.html
            "timeline": CRMService(db).get_unified_timeline(customer_id),
        },
    )


# ── Update ────────────────────────────────────────────────────────────────────

@router.post("/{customer_id}/post-interest", response_class=RedirectResponse)
async def customer_post_interest(
    customer_id: int,
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user_id),
):
    """§21 — charge accrued interest: create a DRAFT finance-charge invoice the
    operator then reviews + finalizes. Redirects to the new draft on success."""
    try:
        inv = CRMService(db, current_user_id=user_id).post_interest_charge(customer_id)
    except ValueError as exc:
        db.rollback()
        return RedirectResponse(
            f"/customers/{customer_id}?error={url_quote(str(exc))}", status_code=303)
    except Exception:
        db.rollback()
        log.exception("Unexpected error posting interest for customer %s", customer_id)
        return RedirectResponse(
            f"/customers/{customer_id}?error={url_quote('Unexpected error — interest was not charged.')}",
            status_code=303)
    return RedirectResponse(f"/invoices/{inv.id}?ok={url_quote('Draft finance-charge invoice created — review and finalize.')}", status_code=303)


@router.post("/{customer_id}", response_class=RedirectResponse)
async def customer_update(
    customer_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user_id),
):
    c = db.query(Customer).filter(Customer.id == customer_id).first()
    if not c:
        return RedirectResponse("/customers/", status_code=303)

    form = await request.form()

    # QA — server-side email-format validation (parity with create). Blank stays
    # allowed; reject a malformed value before mutating anything.
    _new_email = str(form.get("email", "")).strip()
    if _new_email and not is_valid_email(_new_email):
        return RedirectResponse(
            f"/customers/{customer_id}?error="
            + url_quote("Please enter a valid email address (e.g. name@example.com)."),
            status_code=303,
        )

    # C10 — block editing email to one another customer already owns (the unique
    # index would otherwise 500 at commit). Check before mutating anything.
    if _new_email:
        _conflict = _find_customer_by_email(db, _new_email, exclude_id=customer_id)
        if _conflict is not None:
            return RedirectResponse(
                f"/customers/{customer_id}?error="
                + url_quote(f"Another customer already uses that email: {_conflict.company_name}"),
                status_code=303,
            )

    # QA — block editing phone to one another ACTIVE customer already owns
    # (counter staff look people up by phone). Self excluded; blank exempt. Phone
    # is not a unique DB key, so unlike email there's no commit-time backstop —
    # this pre-check is the guard. The operator changes the phone or merges.
    _new_phone = str(form.get("phone", "")).strip()
    if _new_phone:
        _phone_conflict = _find_customer_by_phone(db, _new_phone, exclude_id=customer_id)
        if _phone_conflict is not None:
            return RedirectResponse(
                f"/customers/{customer_id}?error="
                + url_quote(f"Another customer already uses that phone: {_phone_conflict.company_name}"),
                status_code=303,
            )

    c.company_name = str(form.get("company_name", "")).strip()
    c.contact_name = str(form.get("contact_name", "")).strip()
    c.account_number = str(form.get("account_number", "")).strip()
    c.phone = str(form.get("phone", "")).strip()
    c.email = str(form.get("email", "")).strip()
    c.address_line1 = str(form.get("address_line1", "")).strip()
    c.address_line2 = str(form.get("address_line2", "")).strip()
    c.city = str(form.get("city", "")).strip()
    c.state = str(form.get("state", "")).strip()
    c.zip_code = str(form.get("zip_code", "")).strip()
    c.payment_terms = str(form.get("payment_terms", PaymentTerms.COD))
    c.pricing_tier  = str(form.get("pricing_tier", "standard"))
    c.credit_limit  = float(form.get("credit_limit") or 0)
    # Out-of-range standing discount bricks quote/invoice creation for this
    # customer (document services now reject it) — block it at the source.
    try:
        _discount = float(form.get("discount_pct") or 0)
    except (TypeError, ValueError):
        _discount = -1.0
    if not (0.0 <= _discount <= 100.0):
        return RedirectResponse(
            f"/customers/{customer_id}?error="
            + url_quote("Standing discount must be between 0 and 100."),
            status_code=303,
        )
    c.discount_pct  = _discount
    c.interest_rate = float(form.get("interest_rate") or 0)
    _cs = str(form.get("card_surcharge_pct", "")).strip()
    c.card_surcharge_pct = float(_cs) if _cs else None  # blank → NULL = use system default; 0 = no surcharge
    c.is_tax_exempt = bool(form.get("is_tax_exempt"))
    c.tax_exempt_cert_number = str(form.get("tax_exempt_cert_number", "")).strip() or None
    c.notes = str(form.get("notes", "")).strip()
    c.internal_notes = str(form.get("internal_notes", "")).strip()
    # P2-D6 — only touch type when the form sends it (older edit forms don't, and
    # must not silently reset an existing customer's type to OTHER).
    if "customer_type" in form:
        c.customer_type = _form_customer_type(form)
    if "customer_status" in form:
        c.customer_status = _form_customer_status(form)
    # P2-D2 — only rewrite flags when the chip editor submitted them (hidden
    # flags_submitted marker), else a save from a form without the editor would
    # wipe the flags. set_stored_flags leaves tax-exempt / contact-method alone.
    if form.get("flags_submitted"):
        # set_stored_flags syncs customer_status ↔ CREDIT_HOLD (go-live bug #3).
        CustomerService(db).set_stored_flags(c, form.getlist("flags"))
    elif "customer_status" in form:
        # When the status dropdown is saved WITHOUT the chip editor, mirror the
        # CREDIT_HOLD flag in the stored CSV so the two fields stay consistent.
        from app.constants import CustomerStatus as _CS, CustomerFlag as _CF
        _stored = set(CustomerService._parse_stored(c.flags))
        if c.customer_status == _CS.CREDIT_HOLD:
            _stored.add(_CF.CREDIT_HOLD)
        else:
            _stored.discard(_CF.CREDIT_HOLD)
        c.flags = CustomerService._serialize_stored(_stored)
    try:
        db.commit()
    except IntegrityError:
        # C10 race backstop — another insert/edit took the email between the
        # pre-check and commit.
        db.rollback()
        return RedirectResponse(
            f"/customers/{customer_id}?error="
            + url_quote("Could not save — that email is already used by another customer."),
            status_code=303,
        )
    return RedirectResponse(f"/customers/{customer_id}?saved=1", status_code=303)


# ── SMS consent + opt-out controls (§22 Function B) ──────────────────────────
# "OK to text" / "Stop texting" / "Do Not Contact" toggles on the customer
# profile + communications page. Each posts a normal form and redirects 303
# back to the doc it was triggered from (Referer), defaulting to the detail page.
# The shared MessagingService owns the consent/opt-out semantics; these routes
# are thin form handlers around it.

def _consent_redirect(request: Request, customer_id: int) -> str:
    """Redirect target for the consent toggles — back to the referring customer
    page (detail or communications), defaulting to the detail page. Only same-app
    customer URLs are honored so the Referer header can't bounce us off-site."""
    ref = request.headers.get("referer", "") or ""
    if ref:
        from urllib.parse import urlsplit
        path = urlsplit(ref).path
        if path.startswith(f"/customers/{customer_id}"):
            return path
    return f"/customers/{customer_id}"


@router.post("/{customer_id}/sms-consent", response_class=RedirectResponse)
def customer_sms_consent(
    customer_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user_id),
):
    """Record verbal SMS consent — sets allow_sms + sms_consent_at + method.
    Audited + committed inside MessagingService.record_consent."""
    MessagingService(db, current_user_id=user_id).record_consent(
        customer_id, CommunicationChannel.SMS, "verbal"
    )
    return RedirectResponse(_consent_redirect(request, customer_id), status_code=303)


@router.post("/{customer_id}/sms-optout", response_class=RedirectResponse)
def customer_sms_optout(
    customer_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user_id),
):
    """Stop texting this customer — clears allow_sms only (email untouched, and
    they are NOT marked do_not_contact). Audited like the nearby routes."""
    c = db.query(Customer).filter(Customer.id == customer_id).first()
    if c:
        c.allow_sms = False
        MessagingService(db, current_user_id=user_id).audit(
            entity_type="customer",
            entity_id=customer_id,
            action="sms_opt_out",
            new_value={"allow_sms": False},
        )
        db.commit()
    return RedirectResponse(_consent_redirect(request, customer_id), status_code=303)


@router.post("/{customer_id}/contact-optout", response_class=RedirectResponse)
def customer_contact_optout(
    customer_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user_id),
):
    """Mark Do Not Contact — sets do_not_contact and clears allow_sms/allow_email.
    Audited + committed inside MessagingService.record_opt_out."""
    MessagingService(db, current_user_id=user_id).record_opt_out(
        customer_id, reason="manual"
    )
    return RedirectResponse(_consent_redirect(request, customer_id), status_code=303)


@router.post("/{customer_id}/contact-allow", response_class=RedirectResponse)
def customer_contact_allow(
    customer_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user_id),
):
    """Undo Do Not Contact — clears do_not_contact (does not silently re-grant
    SMS consent; the rep re-arms "OK to text" explicitly). Audited."""
    c = db.query(Customer).filter(Customer.id == customer_id).first()
    if c:
        c.do_not_contact = False
        MessagingService(db, current_user_id=user_id).audit(
            entity_type="customer",
            entity_id=customer_id,
            action="contact_allowed",
            new_value={"do_not_contact": False},
        )
        db.commit()
    return RedirectResponse(_consent_redirect(request, customer_id), status_code=303)


# ── Log Call (HTMX) ───────────────────────────────────────────────────────────

@router.post("/{customer_id}/log-call", response_class=HTMLResponse)
async def log_call(
    customer_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user_id),
):
    form = await request.form()
    call_type = str(form.get("call_type", CallType.INBOUND))
    outcome = str(form.get("outcome", CallOutcome.OTHER))
    notes = str(form.get("notes", "")).strip()

    crm = CRMService(db, current_user_id=user_id)
    entry = crm.log_call(
        customer_id=customer_id,
        call_type=call_type,
        outcome=outcome,
        notes=notes,
    )

    return templates.TemplateResponse(
        request,
        "customers/_call_log_row.html",
        {"log": entry},
    )


# ── Communications Timeline ───────────────────────────────────────────────────

_RELATED_ENTITY_HREF = {
    "quote":      "/quotes/{id}",
    "invoice":    "/invoices/{id}",
    "so":         "/sales-orders/{id}",
    "po":         "/purchase-orders/{id}",
    "ra":         "/returns/{id}",
    "warranty":   "/warranty/{id}",
    "research":   "/research/{id}",
    "core_slip":  "/cores/{id}",
    "statement":  "/customers/{id}/statement",
}


@router.get("/{customer_id}/communications", response_class=HTMLResponse)
def customer_communications(
    customer_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Communication timeline for a customer.

    Lists every email, SMS, phone-call note, and manual comm logged to this
    customer — outbound and inbound — newest first. Read-only (immutable log).
    """
    c = db.query(Customer).filter(Customer.id == customer_id).first()
    if c is None:
        return RedirectResponse("/customers/", status_code=303)

    comms = (
        db.query(Communication)
        .filter(Communication.customer_id == customer_id)
        .order_by(Communication.sent_at.desc())
        .limit(500)
        .all()
    )

    # Build a related-doc href lookup per row so the template can stay dumb
    related_links: dict[int, dict[str, str] | None] = {}
    for comm in comms:
        if comm.related_entity_type and comm.related_entity_id:
            tmpl = _RELATED_ENTITY_HREF.get(comm.related_entity_type)
            if tmpl:
                related_links[comm.id] = {
                    "label": f"{comm.related_entity_type.replace('_', ' ').title()} #{comm.related_entity_id}",
                    "href":  tmpl.format(id=comm.related_entity_id),
                }
            else:
                related_links[comm.id] = {
                    "label": f"{comm.related_entity_type} #{comm.related_entity_id}",
                    "href":  "",
                }
        else:
            related_links[comm.id] = None

    return templates.TemplateResponse(
        request,
        "customers/communications.html",
        {
            "request":        request,
            "customer":       c,
            "comms":          comms,
            "related_links":  related_links,
        },
    )


# Maps the form's comm_type select value → (channel, direction)
_COMM_TYPE_MAP: dict[str, tuple[str, str]] = {
    "phone_outbound": (CommunicationChannel.PHONE_CALL, CommunicationDirection.OUTBOUND),
    "phone_inbound":  (CommunicationChannel.PHONE_CALL, CommunicationDirection.INBOUND),
    "email_outbound": (CommunicationChannel.EMAIL,       CommunicationDirection.OUTBOUND),
    "email_inbound":  (CommunicationChannel.EMAIL,       CommunicationDirection.INBOUND),
    "sms_outbound":   (CommunicationChannel.SMS,         CommunicationDirection.OUTBOUND),
    "sms_inbound":    (CommunicationChannel.SMS,         CommunicationDirection.INBOUND),
    "note":           (CommunicationChannel.MANUAL_NOTE, CommunicationDirection.OUTBOUND),
}


@router.post("/{customer_id}/communications/log", response_class=RedirectResponse)
async def customer_communications_log(
    customer_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user_id),
):
    """Log a manual communication entry (no real send — all go through NullProvider)."""
    form = await request.form()
    comm_type      = str(form.get("comm_type", "phone_outbound"))
    body           = str(form.get("body", "")).strip()
    subject        = str(form.get("subject", "")).strip() or None
    contact_addr   = str(form.get("contact_address", "")).strip()
    rel_type       = str(form.get("related_entity_type", "")).strip() or None
    rel_id_raw     = str(form.get("related_entity_id", "")).strip()
    rel_id         = int(rel_id_raw) if rel_id_raw.isdigit() else None

    channel, direction = _COMM_TYPE_MAP.get(
        comm_type,
        (CommunicationChannel.MANUAL_NOTE, CommunicationDirection.OUTBOUND),
    )

    svc = MessagingService(db, current_user_id=user_id)
    if direction == CommunicationDirection.INBOUND:
        svc.record_inbound(
            customer_id=customer_id,
            channel=channel,
            body=body,
            subject=subject,
            from_address=contact_addr,
            related_entity_type=rel_type,
            related_entity_id=rel_id,
        )
    else:
        svc.log_manual_communication(
            customer_id=customer_id,
            channel=channel,
            body=body,
            subject=subject,
            to_address=contact_addr,
            related_entity_type=rel_type,
            related_entity_id=rel_id,
        )

    return RedirectResponse(
        f"/customers/{customer_id}/communications",
        status_code=303,
    )


# ── Balance Widget (HTMX) ─────────────────────────────────────────────────────

@router.get("/{customer_id}/balance", response_class=HTMLResponse)
def customer_balance(
    customer_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user_id),
):
    crm = CRMService(db, current_user_id=user_id)
    balance = crm.get_account_balance(customer_id)
    return templates.TemplateResponse(
        request,
        "customers/_balance_widget.html",
        {"balance": balance},
    )


# ── Deactivate / Reactivate ─────────────────────────────────────────────────

@router.post("/{customer_id}/deactivate", response_class=RedirectResponse)
def customer_deactivate(
    customer_id: int,
    db: Session = Depends(get_db),
):
    c = db.query(Customer).filter(Customer.id == customer_id).first()
    if c:
        c.is_active = False
        db.commit()
    return RedirectResponse("/customers/", status_code=303)


@router.post("/{customer_id}/reactivate", response_class=RedirectResponse)
def customer_reactivate(
    customer_id: int,
    db: Session = Depends(get_db),
):
    """Undo a deactivate — restores the customer to the active lists.
    Redirects back to the detail page (not the list) so the user sees the
    now-reactivated record, mirroring /vendors/{id}/reactivate."""
    c = db.query(Customer).filter(Customer.id == customer_id).first()
    if c:
        c.is_active = True
        db.commit()
    return RedirectResponse(f"/customers/{customer_id}", status_code=303)


# ── Statement ─────────────────────────────────────────────────────────────────

def _optional_user_id(request: Request) -> int | None:
    """Resolve the signed-in user id WITHOUT enforcing auth (R3 — statement
    persistence attribution only; the route itself stays as permissive as it
    was before persistence existed)."""
    try:
        from app.auth import read_session_token, SESSION_COOKIE
        return read_session_token(request.cookies.get(SESSION_COOKIE))
    except Exception:
        return None


@router.post("/statements/bulk-generate", response_class=RedirectResponse)
def customers_bulk_statements(
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user_id),
):
    """§21 — month-end: generate + persist a statement for every customer with an
    open AR balance, for the current calendar month. Redirects to AR aging with a
    summary flash."""
    from datetime import datetime
    from app.services.statement_service import StatementService
    # Use the UTC date — invoices are timestamped UTC (func.now()), so a local
    # date.today() near midnight could exclude "today's" just-created invoices.
    today = datetime.utcnow().date()
    period_start = today.replace(day=1)
    try:
        summary = StatementService(db).generate_bulk_statements(
            period_start, today, generated_by_user_id=user_id)
    except Exception:
        log.exception("bulk statement generation failed")
        return RedirectResponse(
            f"/reports/ar-aging?error={url_quote('Bulk statement generation failed.')}",
            status_code=303)
    msg = f"Generated {summary['generated']} statement(s) for {summary['customers']} customer(s) with a balance."
    return RedirectResponse(f"/reports/ar-aging?ok={url_quote(msg)}", status_code=303)


@router.get("/{customer_id}/statement", response_class=HTMLResponse)
def customer_statement_form(
    customer_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Statement date-range selector page + saved-statement history (R3)."""
    from datetime import date
    from app.services.statement_service import StatementService
    c = db.query(Customer).filter(Customer.id == customer_id).first()
    if c is None:
        return RedirectResponse("/customers/", status_code=303)
    today = date.today()
    # Default: first of current month → today
    default_start = today.replace(day=1)
    return templates.TemplateResponse(
        request,
        "customers/statement_form.html",
        {
            "customer": c,
            "default_start": default_start.isoformat(),
            "default_end": today.isoformat(),
            # R3 — archived statements (immutable snapshots of what was sent)
            "history": StatementService(db).get_statement_history(customer_id),
        },
    )


@router.get("/{customer_id}/statement/print", response_class=HTMLResponse)
def customer_statement_print(
    customer_id: int,
    request: Request,
    start: str = "",
    end: str = "",
    db: Session = Depends(get_db),
):
    """Print-ready statement HTML. Also used by /pdf to generate the PDF bytes."""
    from datetime import date
    from app.services.statement_service import StatementService
    from app.settings_utils import get_setting_value_db

    today = date.today()
    try:
        period_start = date.fromisoformat(start) if start else today.replace(day=1)
        period_end = date.fromisoformat(end) if end else today
    except ValueError:
        period_start = today.replace(day=1)
        period_end = today

    svc = StatementService(db)
    stmt = svc.generate_statement(
        customer_id=customer_id,
        period_start=period_start,
        period_end=period_end,
    )
    # R3 — the print view is the "this went to the customer" moment: persist an
    # immutable snapshot (idempotent per customer+period+day, see service).
    saved = svc.persist_statement(stmt, generated_by_user_id=_optional_user_id(request))
    company = {
        "name":    get_setting_value_db(db, "company_name",    "JAKS Parts"),
        "address": get_setting_value_db(db, "company_address", ""),
        "phone":   get_setting_value_db(db, "company_phone",   ""),
        "email":   get_setting_value_db(db, "company_email",   ""),
    }
    return templates.TemplateResponse(
        request,
        "customers/statement_print.html",
        {"stmt": stmt, "company": company,
         "statement_number": saved.statement_number},
    )


@router.get("/{customer_id}/statement/pdf")
def customer_statement_pdf(
    customer_id: int,
    request: Request,
    start: str = "",
    end: str = "",
    db: Session = Depends(get_db),
):
    """Server-side PDF via WeasyPrint. Falls back to print view if GTK missing."""
    from datetime import date
    from urllib.parse import quote as url_quote
    from app.services.statement_service import StatementService
    from app.settings_utils import get_setting_value_db
    from fastapi.responses import Response as FastAPIResponse

    today = date.today()
    try:
        period_start = date.fromisoformat(start) if start else today.replace(day=1)
        period_end = date.fromisoformat(end) if end else today
    except ValueError:
        period_start = today.replace(day=1)
        period_end = today

    svc = StatementService(db)
    stmt = svc.generate_statement(
        customer_id=customer_id,
        period_start=period_start,
        period_end=period_end,
    )
    # R3 — PDF generation is also a "sent to customer" moment; same idempotent
    # persistence as /print (same customer+period+day reuses the row).
    saved = svc.persist_statement(stmt, generated_by_user_id=_optional_user_id(request))
    company = {
        "name":    get_setting_value_db(db, "company_name",    "JAKS Parts"),
        "address": get_setting_value_db(db, "company_address", ""),
        "phone":   get_setting_value_db(db, "company_phone",   ""),
        "email":   get_setting_value_db(db, "company_email",   ""),
    }
    html_str = templates.env.get_template("customers/statement_print.html").render(
        request=request, stmt=stmt, company=company,
        statement_number=saved.statement_number,
    )
    try:
        from weasyprint import HTML
        from app.services.document_render import static_url_fetcher
        pdf_bytes = HTML(
            string=html_str, base_url=str(request.base_url),
            url_fetcher=static_url_fetcher,
        ).write_pdf()
    except Exception:
        return RedirectResponse(
            f"/customers/{customer_id}/statement/print?start={start}&end={end}",
            status_code=302,
        )
    safe_name = url_quote(stmt["customer"].company_name.replace("/", "-"))
    filename = f"Statement_{safe_name}_{period_end.isoformat()}.pdf"
    return FastAPIResponse(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": str(len(pdf_bytes)),
        },
    )


@router.get("/{customer_id}/statement/archive/{statement_id}", response_class=HTMLResponse)
def customer_statement_archive(
    customer_id: int,
    statement_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """R3 — read-only ARCHIVED statement, re-rendered from snapshot_json.

    This is the dispute-resolution view: it shows exactly what was generated
    and sent, never live data. No persistence, no recalculation."""
    from app.models.statement import CustomerStatement
    from app.services.statement_service import StatementService
    from app.settings_utils import get_setting_value_db

    row = (
        db.query(CustomerStatement)
        .filter(
            CustomerStatement.id == statement_id,
            CustomerStatement.customer_id == customer_id,
        )
        .first()
    )
    if row is None:
        return RedirectResponse(f"/customers/{customer_id}/statement", status_code=303)

    stmt = StatementService.snapshot_to_render_ctx(row)
    if stmt is None:  # legacy row persisted before snapshots existed
        return RedirectResponse(f"/customers/{customer_id}/statement", status_code=303)

    company = {
        "name":    get_setting_value_db(db, "company_name",    "JAKS Parts"),
        "address": get_setting_value_db(db, "company_address", ""),
        "phone":   get_setting_value_db(db, "company_phone",   ""),
        "email":   get_setting_value_db(db, "company_email",   ""),
    }
    return templates.TemplateResponse(
        request,
        "customers/statement_print.html",
        {
            "stmt": stmt,
            "company": company,
            "statement_number": row.statement_number,
            "archived": True,
            "archived_generated_at": row.generated_at,
        },
    )
